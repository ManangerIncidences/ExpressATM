from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, JSONResponse
import logging
import sqlite3
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional, Dict
from datetime import date, datetime
import io
from ..database import get_db
from ..models import Alert, SalesRecord, Agency, MonitoringSession, SystemLog
from ..alerts import AlertSystem
from ..agency_behavior_analyzer import agency_analyzer, is_agency_unusual
import numpy as np
from ..dom_learning_engine import dom_learner

# 🔗 IMPORTACIÓN INTELIGENTE DE SCHEDULER
try:
    from ..scheduler_hybrid import monitoring_scheduler
    USING_HYBRID_SCHEDULER = True
    print("Usando scheduler hibrido")
except ImportError:
    from ..scheduler import monitoring_scheduler
    USING_HYBRID_SCHEDULER = False
    print("Usando scheduler clasico")

from ..intelligence import intelligence_engine
from pydantic import BaseModel
# Nota: se eliminó el import redundante de monitoring_scheduler para no sobreescribir
# la instancia híbrida cuando está disponible.

router = APIRouter()
logger = logging.getLogger(__name__)

# Modelos Pydantic para respuestas
class AlertResponse(BaseModel):
    id: int
    agency_code: str
    agency_name: str
    alert_type: str
    alert_message: str
    current_sales: float
    current_balance: float
    is_reported: bool
    alert_date: datetime
    reported_at: Optional[datetime] = None
    
    class Config:
        from_attributes = True

class AgencyResponse(BaseModel):
    code: str
    name: str
    latest_sales: Optional[float] = None
    latest_balance: Optional[float] = None
    latest_update: Optional[datetime] = None

class MonitoringStatusResponse(BaseModel):
    is_running: bool
    current_session_id: Optional[int] = None
    monitoring_interval: int
    session_start: Optional[str] = None
    total_iterations: Optional[int] = None
    total_agencies_processed: Optional[int] = None
    total_alerts_generated: Optional[int] = None
    next_run_time: Optional[str] = None

class ReportAlertRequest(BaseModel):
    alert_id: int

# Nuevos modelos para Inteligencia Artificial
class IntelligenceStatusResponse(BaseModel):
    intelligence_enabled: bool
    adaptive_config: Optional[dict] = None
    failure_prediction: Optional[dict] = None
    anomalies_detected: Optional[int] = None
    optimizations_available: Optional[int] = None
    system_metrics: Optional[dict] = None
    error: Optional[str] = None

class OptimizationResponse(BaseModel):
    parameter: str
    current_value: str
    recommended_value: str
    confidence: float
    reason: str
    expected_improvement: str

class PredictionResponse(BaseModel):
    prediction_type: str
    probability: float
    confidence: float
    recommendation: str
    details: dict

class IterationProgressResponse(BaseModel):
    active: bool
    steps: list
    started_at: Optional[str] = None
    updated_at: Optional[str] = None
    finished_at: Optional[str] = None
    current: Optional[str] = None
    error: Optional[str] = None
    version: Optional[int] = None

class IntelligenceToggleRequest(BaseModel):
    enabled: bool

class AdaptiveConfigRequest(BaseModel):
    config_updates: dict

# =============================================================
# Exportación de incidencias (alertas) - CSV / Excel / PDF
# =============================================================
def _query_alerts_for_export(db: Session, start_day: str | None, end_day: str | None, reported: Optional[bool], alert_type: Optional[str]):
    q = db.query(Alert)
    if start_day:
        try:
            datetime.fromisoformat(start_day)
            q = q.filter(Alert.alert_day >= start_day)
        except Exception:
            pass
    if end_day:
        try:
            datetime.fromisoformat(end_day)
            q = q.filter(Alert.alert_day <= end_day)
        except Exception:
            pass
    if reported is not None:
        q = q.filter(Alert.is_reported == reported)
    if alert_type:
        q = q.filter(Alert.alert_type == alert_type)
    return q.order_by(Alert.alert_day.desc(), Alert.created_at.desc()).all()

@router.get("/alerts/export")
async def export_alerts(
    format: str = Query("csv", description="csv|excel|pdf"),
    start_day: Optional[str] = Query(None, description="YYYY-MM-DD desde"),
    end_day: Optional[str] = Query(None, description="YYYY-MM-DD hasta"),
    reported: Optional[bool] = Query(None, description="Filtrar por reportadas (true/false)"),
    alert_type: Optional[str] = Query(None, description="Tipo de alerta"),
    db: Session = Depends(get_db)
):
    alerts = _query_alerts_for_export(db, start_day, end_day, reported, alert_type)
    if format not in {"csv", "excel", "pdf"}:
        raise HTTPException(status_code=400, detail="Formato no soportado")
    if not alerts:
        return JSONResponse({"message": "Sin alertas para el rango indicado"})

    # Datos base
    rows = []
    for a in alerts:
        rows.append({
            "id": a.id,
            "fecha": a.alert_day,
            "hora": a.alert_date.strftime('%H:%M:%S') if a.alert_date else '',
            "agencia_codigo": a.agency_code,
            "agencia_nombre": a.agency_name,
            "sorteo": a.lottery_type or '',
            "tipo": a.alert_type,
            "mensaje": a.alert_message,
            "ventas": a.current_sales,
            "balance": a.current_balance,
            "reportada": a.is_reported,
            "reportada_at": a.reported_at.strftime('%Y-%m-%d %H:%M:%S') if a.reported_at else ''
        })

    if format == "csv":
        import csv
        # Orden explícito de columnas CSV
        csv_fields = ["hora", "agencia_nombre", "sorteo", "tipo", "ventas", "balance", "reportada", "reportada_at"]
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=csv_fields)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, '') for k in csv_fields})
        buffer.seek(0)
        return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={
            "Content-Disposition": f"attachment; filename=alertas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        })
    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
        from openpyxl.utils import get_column_letter

        type_map = {
            'sustained_growth': 'Crec. Sostenido',
            'growth_variation': 'Var. Crecimiento',
            'threshold': 'Umbral'
        }
        # Texto rango
        if start_day and end_day:
            if start_day == end_day:
                rango_txt = f"Fecha: {start_day}"
            else:
                rango_txt = f"Desde: {start_day}   Hasta: {end_day}"
        elif start_day and not end_day:
            rango_txt = f"Desde: {start_day}"
        elif end_day and not start_day:
            rango_txt = f"Hasta: {end_day}"
        else:
            try:
                dias = sorted({str(r['fecha']) for r in rows})
                if dias:
                    if dias[0] == dias[-1]:
                        rango_txt = f"Fecha: {dias[0]}"
                    else:
                        rango_txt = f"Desde: {dias[0]}   Hasta: {dias[-1]}"
                else:
                    rango_txt = "Fecha: (sin datos)"
            except Exception:
                rango_txt = "Fecha: (n/d)"

        wb = Workbook()
        ws = wb.active
        ws.title = "Alertas"
        ws['A1'] = "ExpressATM"; ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws['A2'] = "Reporte de alertas"; ws['A2'].font = Font(bold=False, size=10, color="464646")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        ws['A3'] = rango_txt; ws['A3'].font = Font(size=9)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
        ws.append([None])
        header_row = 5
        headers_excel = ["Hora", "Agencia", "Sorteo", "Tipo", "Ventas", "Balance", "Hora Reportada"]
        ws.append(headers_excel)
        for r in rows:
            hora_rep = None
            if r.get('reportada_at'):
                try:
                    hora_rep = r['reportada_at'][11:19]
                except Exception:
                    hora_rep = r['reportada_at']
            ws.append([
                r['hora'], r['agencia_nombre'], r.get('sorteo') or '',
                type_map.get(r['tipo'], r['tipo']), r['ventas'], r['balance'], hora_rep or 'null'
            ])
        thin = Side(style="thin", color="AAAAAA")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill_header = PatternFill("solid", fgColor="F2F2F2")
        for c, title in enumerate(headers_excel, 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_header
            cell.border = border
        ventas_col = headers_excel.index("Ventas") + 1
        balance_col = headers_excel.index("Balance") + 1
        for r_idx in range(header_row + 1, ws.max_row + 1):
            for c_idx in range(1, len(headers_excel) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = border
                header_name = headers_excel[c_idx-1]
                if header_name in ("Tipo", "Hora Reportada", "Sorteo"):
                    cell.alignment = Alignment(horizontal='center')
                if c_idx in (ventas_col, balance_col) and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$" #,##0'
                    if c_idx == balance_col and cell.value < 0:
                        cell.font = Font(color="C00000")
        # Ajuste: ampliar 'Sorteo' para que quepa CHANCE_EXPRESS / RULETA_EXPRESS
        widths = [10, 50, 22, 18, 12, 12, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row_idx in (1, 2, 3):
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
            "Content-Disposition": f"attachment; filename=alertas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        })
    else:  # pdf
        try:
            from fpdf import FPDF
        except ImportError:
            raise HTTPException(status_code=500, detail="Dependencia fpdf2 no instalada")
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        # Logo (si existe) – se intenta cargar sin romper exporte
        try:
            import os
            # Prioridad: imag_logo principal para contenido; logo.png solo fallback
            logo_paths = [
                'frontend/logos/imag_logo.png',
                'logos/imag_logo.png',
                'frontend/logos/logo.png',
                'logos/logo.png'
            ]
            for _lp in logo_paths:
                if os.path.isfile(_lp):
                    pdf.image(_lp, x=10, y=8, h=10)
                    pdf.set_xy(19, 10)
                    break
        except Exception:
            pass
        # Título y subtítulo refinados
        # Reqs:
        # 1) Más espacio entre título y subtítulo.
        # 2) Reducir título al tamaño previo del subtítulo.
        # 3) Subtítulo más pequeño y sin negritas.
        # 4) Diferencia clara (peso, color suave y espaciado).
        pdf.set_font("Arial", 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 7, "ExpressATM", ln=1, align='L')
        pdf.ln(2)
        pdf.set_font("Arial", '', 10)
        pdf.set_text_color(70, 70, 70)
        pdf.cell(0, 5, "Reporte de alertas", ln=1, align='L')
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)
        # Determinar rango mostrado con formato solicitado (Desde/Hasta o Fecha única)
        if start_day and end_day:
            if start_day == end_day:
                rango_txt = f"Fecha: {start_day}"
            else:
                rango_txt = f"Desde: {start_day}   Hasta: {end_day}"
        elif start_day and not end_day:
            rango_txt = f"Desde: {start_day}"
        elif end_day and not start_day:
            rango_txt = f"Hasta: {end_day}"
        else:
            try:
                dias = sorted({str(r['fecha']) for r in rows})
                if dias:
                    if dias[0] == dias[-1]:
                        rango_txt = f"Fecha: {dias[0]}"
                    else:
                        rango_txt = f"Desde: {dias[0]}   Hasta: {dias[-1]}"
                else:
                    rango_txt = "Fecha: (sin datos)"
            except Exception:
                rango_txt = "Fecha: (n/d)"
        pdf.set_font("Arial", '', 9)
        pdf.cell(0, 6, rango_txt, ln=1, align='R')
        pdf.ln(1)
        pdf.set_font("Arial", 'B', 8)
        # Nuevos requisitos:
        # 1) Tabla ocupa todo el ancho -> redistribuimos anchos hasta ~273mm (margen estándar 10mm cada lado en A4 apaisado)
        # 2) Remover columna Fecha.
        # 3) Columna Hora = hora alerta.
        # 4) Columna Agencia (antes Nombre).
        # 5) Columna Tipo traducida español.
        # 6) Columna Rep. ahora muestra hora de reporte (o 'null').
        headers = ["Hora", "Agencia", "Sorteo", "Tipo", "Ventas", "Balance", "Hora Reportada"]
        # Ajuste de anchos: ampliar 'Sorteo' y compensar reduciendo ligeramente otras columnas
        col_widths = [15, 100, 30, 38, 22, 22, 35]
        # Borde y estilo suave
        pdf.set_draw_color(140, 140, 140)
        pdf.set_line_width(0.1)
        for idx, (h, w) in enumerate(zip(headers, col_widths)):
            align = 'C' if h in {"Tipo", "Ventas", "Balance", "Hora Reportada"} else 'L'
            pdf.cell(w, 7, h, border=1, align=align)
        pdf.ln()
        pdf.set_font("Arial", size=7)
        # Mapeo de tipos a español
        type_map = {
            'sustained_growth': 'Crec. Sostenido',
            'growth_variation': 'Var. Crecimiento',
            'threshold': 'Umbral'
        }
        # Filas
        def _safe(txt: str) -> str:
            if not txt:
                return ''
            # Reemplazar caracteres fuera de Latin-1 (como emojis) por '?'
            return ''.join(c if ord(c) < 256 else '?' for c in str(txt))
        def _fmt_currency(val):
            if val is None:
                return ''
            try:
                n = int(round(float(val)))
            except Exception:
                return str(val)
            # Formato español sin decimales: separador miles '.'
            s = f"{abs(n):,}".replace(',', '.')
            s = f"$ {s}" if n >= 0 else f"-$ {s}"  # signo antes
            return s
        def _truncate(txt: str, max_w: float) -> str:
            if not txt:
                return ''
            # Intentamos truncar usando '...' ASCII para evitar caracteres fuera Latin-1
            ellipsis = '...'
            if pdf.get_string_width(txt) <= max_w:
                return txt
            # Reducir hasta que quepa con '...'
            base = txt
            while len(base) > 1 and pdf.get_string_width(base + ellipsis) > max_w:
                base = base[:-1]
            return (base + ellipsis) if base else ''

        for r in rows[:800]:  # límite precautorio
            hora_alerta = r['hora']
            agencia = _truncate(_safe((r['agencia_nombre'] or '')[:120]), col_widths[1]-1)
            sorteo_val = _truncate(_safe((r.get('sorteo') or '')[:20]), col_widths[2]-1)
            tipo = _truncate(_safe(type_map.get(r['tipo'], r['tipo'])), col_widths[3]-1)
            ventas = _fmt_currency(r['ventas']) if r['ventas'] is not None else ''
            balance = _fmt_currency(r['balance']) if r['balance'] is not None else ''
            hora_reporte = 'null'
            if r.get('reportada_at'):
                try:
                    hora_reporte = r['reportada_at'][11:19]
                except Exception:
                    hora_reporte = r['reportada_at']
            cells = [hora_alerta, agencia, sorteo_val, tipo, ventas, balance, hora_reporte]
            align_map = ['L', 'L', 'C', 'C', 'C', 'C', 'C']
            for (txt, w, align, colname) in zip(cells, col_widths, align_map, headers):
                # Color rojo para balance negativo
                if colname == 'Balance' and r['balance'] is not None and r['balance'] < 0:
                    pdf.set_text_color(200, 0, 0)
                else:
                    pdf.set_text_color(0, 0, 0)
                safe_txt = _safe(txt)
                if pdf.get_string_width(safe_txt) > (w - 1):
                    safe_txt = _truncate(safe_txt, w - 1)
                pdf.cell(w, 5, safe_txt, border=1, align=align)
            pdf.ln()
        # Reset color texto
        pdf.set_text_color(0,0,0)
        # Salida PDF: en fpdf2 output(dest='S') devuelve un bytearray (ya bytes listo para enviar).
        # En la versión clásica de pyfpdf devolvía str (latin-1). Hacemos manejo seguro.
        _raw_pdf = pdf.output(dest='S')
        if isinstance(_raw_pdf, (bytes, bytearray)):
            pdf_bytes = bytes(_raw_pdf)
        else:  # fallback string
            pdf_bytes = str(_raw_pdf).encode('latin-1', 'ignore')
        return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers={
            "Content-Disposition": f"attachment; filename=alertas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        })

# Endpoints principales

def _bucket_dt(dt: datetime, granularity: str) -> str:
    """Redondea hacia abajo un datetime a un bucket (string HH:MM) según granularidad."""
    if granularity == '15m':
        minute = (dt.minute // 15) * 15
    elif granularity == '5m':
        minute = (dt.minute // 5) * 5
    else:  # 'iteration' o '1m'
        minute = dt.minute
    return dt.replace(second=0, microsecond=0, minute=minute).strftime('%H:%M')

def _daterange(end_date: date, days: int):
    from datetime import timedelta
    for i in range(1, days+1):
        yield (end_date - timedelta(days=i)).isoformat()

@router.get("/stats/global")
async def get_global_stats(
    day: Optional[str] = Query(None, description="Día YYYY-MM-DD; por defecto hoy"),
    granularity: str = Query('5m', description="iteration|5m|15m"),
    days: int = Query(7, description="Días hacia atrás para baseline"),
    align_dow: bool = Query(True, description="Baseline por mismo día de semana"),
    lottery: Optional[str] = Query(None, description="Filtrar por tipo de lotería (CHANCE_EXPRESS/RULETA_EXPRESS)"),
    db: Session = Depends(get_db)
):
    """Serie global (hoy) agregada por bucket y baseline promedio de días previos."""
    from collections import defaultdict
    import math

    today_str = day or date.today().isoformat()

    # Obtener registros de HOY
    q_today = db.query(SalesRecord).filter(SalesRecord.capture_day == today_str)
    if lottery:
        q_today = q_today.filter(SalesRecord.lottery_type == lottery)
    recs_today = q_today.order_by(SalesRecord.iteration_time.asc()).all()

    buckets_today = defaultdict(lambda: {"sales": 0.0, "balance": 0.0})
    for r in recs_today:
        label = _bucket_dt(r.iteration_time, granularity)
        buckets_today[label]["sales"] += float(r.sales or 0)
        buckets_today[label]["balance"] += float(r.balance or 0)

    labels_sorted = sorted(buckets_today.keys())
    today_sales = [buckets_today[l]["sales"] for l in labels_sorted]
    today_balance = [buckets_today[l]["balance"] for l in labels_sorted]
    # Apuestas = Δ ventas entre buckets
    today_bets = [0.0]
    for i in range(1, len(today_sales)):
        today_bets.append(today_sales[i] - today_sales[i-1])

    # BASELINE: días previos
    # Filtrar días por align_dow si se solicita
    target_dow = None
    if align_dow and recs_today:
        target_dow = recs_today[0].iteration_time.weekday()

    baseline_series = defaultdict(list)  # label -> [totals por día]
    count_days = 0
    for prev_day in _daterange(date.fromisoformat(today_str), days):
        # Si align_dow, validar dow
        if align_dow:
            # buscar un registro cualquiera para obtener dow de ese día
            any_rec = db.query(SalesRecord).filter(SalesRecord.capture_day == prev_day).first()
            if not any_rec:
                continue
            if any_rec.iteration_time.weekday() != target_dow:
                continue
        q_prev = db.query(SalesRecord).filter(SalesRecord.capture_day == prev_day)
        if lottery:
            q_prev = q_prev.filter(SalesRecord.lottery_type == lottery)
        recs_prev = q_prev.all()
        if not recs_prev:
            continue
        tmp = defaultdict(float)
        for r in recs_prev:
            label = _bucket_dt(r.iteration_time, granularity)
            tmp[label] += float(r.sales or 0)
        # Asegurar alineación solo a labels de hoy si existen, si no, todas
        keys = labels_sorted if labels_sorted else sorted(tmp.keys())
        for l in keys:
            baseline_series[l].append(tmp.get(l, 0.0))
        count_days += 1

    baseline_mean = []
    baseline_std = []
    for l in (labels_sorted if labels_sorted else sorted(baseline_series.keys())):
        vals = baseline_series.get(l, [])
        if not vals:
            baseline_mean.append(None)
            baseline_std.append(None)
        else:
            m = float(sum(vals) / len(vals))
            v = float(sum((x - m) ** 2 for x in vals) / len(vals))
            baseline_mean.append(m)
            baseline_std.append(math.sqrt(v))

    # KPIs comparativos
    total_today_sales = sum(today_sales) if today_sales else 0.0
    total_baseline_sales = sum([x for x in baseline_mean if isinstance(x, (int, float))]) if baseline_mean else 0.0
    deviation_pct = ((total_today_sales - total_baseline_sales) / total_baseline_sales * 100.0) if total_baseline_sales else None
    iterations_today = len(labels_sorted)
    iterations_expected = len([x for x in baseline_mean if x is not None]) if baseline_mean else 0
    peak_idx_today = int(today_sales.index(max(today_sales))) if today_sales else None
    peak_hour_today = labels_sorted[peak_idx_today] if peak_idx_today is not None else None
    peak_idx_baseline = int(baseline_mean.index(max([x for x in baseline_mean if x is not None]))) if baseline_mean and any(x is not None for x in baseline_mean) else None
    peak_hour_baseline = (labels_sorted[peak_idx_baseline] if labels_sorted and peak_idx_baseline is not None else None)

    return {
        "labels": labels_sorted,
        "today": {"sales": today_sales, "balance": today_balance, "bets": today_bets},
        "baseline": {"mean_sales": baseline_mean, "std_sales": baseline_std},
        "summary": {
            "total_today_sales": total_today_sales,
            "total_baseline_sales": total_baseline_sales,
            "deviation_sales_pct": deviation_pct,
            "iterations_today": iterations_today,
            "iterations_expected": iterations_expected,
            "peak_hour_today": peak_hour_today,
            "peak_hour_baseline": peak_hour_baseline,
            "baseline_days_used": count_days
        }
    }

@router.get("/alerts", response_model=List[AlertResponse])
async def get_alerts(
    today_only: bool = Query(True, description="Solo alertas de hoy"),
    reported: Optional[bool] = Query(None, description="Filtrar por estado reportado"),
    alert_type: Optional[str] = Query(None, description="Filtrar por tipo de alerta"),
    db: Session = Depends(get_db)
):
    """Obtener alertas con filtros opcionales"""
    query = db.query(Alert)
    
    if today_only:
        today = date.today().isoformat()
        query = query.filter(Alert.alert_day == today)
    
    if reported is not None:
        query = query.filter(Alert.is_reported == reported)
    
    if alert_type:
        query = query.filter(Alert.alert_type == alert_type)
    
    alerts = query.order_by(Alert.created_at.desc()).all()
    return alerts

@router.post("/alerts/{alert_id}/report")
async def report_alert(alert_id: int, db: Session = Depends(get_db)):
    """Marcar una alerta como reportada"""
    alert_system = AlertSystem()
    success = alert_system.mark_alert_as_reported(alert_id, db)
    
    if success:
        return {"message": "Alerta marcada como reportada", "alert_id": alert_id}
    else:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")

@router.post("/alerts/{alert_id}/unreport")
async def unreport_alert(alert_id: int, db: Session = Depends(get_db)):
    """Revertir estado reportado de una alerta (solo el mismo día)."""
    from ..models import Alert
    try:
        alert = db.query(Alert).filter(Alert.id == alert_id).first()
        if not alert:
            raise HTTPException(status_code=404, detail="Alerta no encontrada")
        # Solo permitir si es del día actual
        today = date.today().isoformat()
        if alert.alert_day != today:
            raise HTTPException(status_code=400, detail="Solo se puede desmarcar alertas del día actual")
        alert.is_reported = False
        alert.reported_at = None
        db.commit()
        return {"message": "Alerta revertida", "alert_id": alert_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error revirtiendo alerta")

@router.get("/agencies", response_model=List[AgencyResponse])
async def get_agencies(
    with_alerts_only: bool = Query(False, description="Solo agencias con alertas"),
    today_only: bool = Query(True, description="Solo agencias monitoreadas hoy"),
    db: Session = Depends(get_db)
):
    """Obtener lista de agencias"""
    today = date.today().isoformat()
    
    if with_alerts_only:
        # Obtener agencias que tienen alertas pendientes hoy
        agencies_with_alerts = db.query(Alert.agency_code).filter(
            Alert.alert_day == today,
            Alert.is_reported == False
        ).distinct().all()
        
        agency_codes = [a.agency_code for a in agencies_with_alerts]
        agencies = db.query(Agency).filter(Agency.code.in_(agency_codes)).all()
    else:
        if today_only:
            # Obtener solo agencias que fueron monitoreadas HOY
            agencies_monitored_today = db.query(SalesRecord.agency_code).filter(
                SalesRecord.capture_day == today
            ).distinct().all()
            
            agency_codes = [a.agency_code for a in agencies_monitored_today]
            agencies = db.query(Agency).filter(Agency.code.in_(agency_codes)).all()
        else:
            # Obtener todas las agencias (comportamiento anterior)
            agencies = db.query(Agency).all()
    
    # Enriquecer con datos más recientes DEL DÍA DE HOY
    enriched_agencies = []
    for agency in agencies:
        if today_only:
            # Buscar el registro más reciente DE HOY
            latest_record = db.query(SalesRecord).filter(
                SalesRecord.agency_code == agency.code,
                SalesRecord.capture_day == today
            ).order_by(SalesRecord.iteration_time.desc()).first()
        else:
            # Buscar el registro más reciente de cualquier día
            latest_record = db.query(SalesRecord).filter(
                SalesRecord.agency_code == agency.code
            ).order_by(SalesRecord.iteration_time.desc()).first()
        
        agency_data = AgencyResponse(
            code=agency.code,
            name=agency.name,
            latest_sales=latest_record.sales if latest_record else None,
            latest_balance=latest_record.balance if latest_record else None,
            latest_update=latest_record.iteration_time if latest_record else None
        )
        enriched_agencies.append(agency_data)
    
    return enriched_agencies

@router.get("/agencies/{agency_code}/history")
async def get_agency_history(
    agency_code: str,
    days: int = Query(7, description="Número de días de historial"),
    db: Session = Depends(get_db)
):
    """Obtener historial de una agencia específica"""
    alert_system = AlertSystem()
    history = alert_system.get_agency_growth_history(agency_code, db, days)
    
    if not history["growth_history"]:
        raise HTTPException(status_code=404, detail="No se encontraron datos para esta agencia")
    
    return history

@router.get("/agencies/{agency_code}/day/{day}/iterations")
async def get_agency_day_iterations(
    agency_code: str,
    day: str,  # formato YYYY-MM-DD
    db: Session = Depends(get_db)
):
    """Obtener detalle de iteraciones (movimientos) de una agencia en un día específico.
    Incluye cambios entre iteraciones y alertas ocurridas ese día.
    """
    # Validar formato simple YYYY-MM-DD
    try:
        datetime.fromisoformat(day)
    except Exception:
        raise HTTPException(status_code=400, detail="Formato de día inválido. Use YYYY-MM-DD")

    # Iteraciones del día
    records = db.query(SalesRecord).filter(
        SalesRecord.agency_code == agency_code,
        SalesRecord.capture_day == day
    ).order_by(SalesRecord.iteration_time.asc()).all()

    if not records:
        return {
            "agency_code": agency_code,
            "day": day,
            "iterations": [],
            "alerts": [],
            "total_iterations": 0
        }

    iterations = []
    prev_sales = None
    prev_balance = None
    for r in records:
        delta_sales = (r.sales - prev_sales) if prev_sales is not None else 0
        delta_balance = (r.balance - prev_balance) if prev_balance is not None else 0
        iterations.append({
            "time": r.iteration_time.isoformat(),
            "sales": float(r.sales) if r.sales is not None else 0,
            "balance": float(r.balance) if r.balance is not None else 0,
            "prizes": float(r.prizes) if r.prizes is not None else 0,
            "prizes_paid": float(r.prizes_paid) if r.prizes_paid is not None else 0,
            "lottery_type": r.lottery_type,
            "delta_sales": float(delta_sales),
            "delta_balance": float(delta_balance)
        })
        prev_sales = r.sales
        prev_balance = r.balance

    # Alertas del día
    day_alerts = db.query(Alert).filter(
        Alert.agency_code == agency_code,
        Alert.alert_day == day
    ).order_by(Alert.created_at.asc()).all()

    alerts_payload = [
        {
            "id": a.id,
            "type": a.alert_type,
            "message": a.alert_message,
            "current_sales": float(a.current_sales) if a.current_sales is not None else None,
            "current_balance": float(a.current_balance) if a.current_balance is not None else None,
            "lottery_type": a.lottery_type,
            "created_at": a.created_at.isoformat()
        }
        for a in day_alerts
    ]

    return {
        "agency_code": agency_code,
        "agency_name": records[0].agency_name,
        "day": day,
        "total_iterations": len(iterations),
        "first_time": iterations[0]["time"],
        "last_time": iterations[-1]["time"],
        "iterations": iterations,
        "alerts": alerts_payload
    }

@router.get("/dashboard")
async def get_dashboard_data(db: Session = Depends(get_db)):
    """Obtener datos para el dashboard principal"""
    today = date.today().isoformat()
    
    # Alertas pendientes por tipo
    alerts_by_type = db.query(Alert.alert_type, func.count(Alert.id)).filter(
        Alert.alert_day == today,
        Alert.is_reported == False
    ).group_by(Alert.alert_type).all()
    
    # Total de agencias monitoreadas hoy
    total_agencies_today = db.query(SalesRecord.agency_code).filter(
        SalesRecord.capture_day == today
    ).distinct().count()
    
    # Últimas iteraciones
    latest_records = db.query(SalesRecord).filter(
        SalesRecord.capture_day == today
    ).order_by(SalesRecord.iteration_time.desc()).limit(10).all()
    
    # Estado del monitoreo
    monitoring_status = monitoring_scheduler.get_status()
    
    # Estadísticas de alertas
    total_alerts_today = db.query(Alert).filter(Alert.alert_day == today).count()
    reported_alerts_today = db.query(Alert).filter(
        Alert.alert_day == today,
        Alert.is_reported == True
    ).count()
    
    return {
        "monitoring_status": monitoring_status,
        "alerts_summary": {
            "total_today": total_alerts_today,
            "pending": total_alerts_today - reported_alerts_today,
            "reported": reported_alerts_today,
            "by_type": dict(alerts_by_type)
        },
        "agencies_summary": {
            "total_monitored_today": total_agencies_today
        },
        "latest_activity": [
            {
                "agency_code": record.agency_code,
                "agency_name": record.agency_name,
                "sales": record.sales,
                "balance": record.balance,
                "time": record.iteration_time.isoformat()
            }
            for record in latest_records
        ]
    }

@router.get("/monitoring/status", response_model=MonitoringStatusResponse)
async def get_monitoring_status():
    """Obtener estado del monitoreo"""
    status = monitoring_scheduler.get_status()
    return MonitoringStatusResponse(**status)

@router.get("/monitoring/progress", response_model=IterationProgressResponse)
async def get_iteration_progress():
    """Obtener progreso detallado de la iteración actual (pasos del scraping)."""
    try:
        progress = monitoring_scheduler.get_progress()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo progreso: {e}")
    return IterationProgressResponse(**progress)

@router.post("/monitoring/start")
async def start_monitoring():
    """Iniciar el monitoreo automático"""
    success = monitoring_scheduler.start_monitoring()
    if success:
        return {"message": "Monitoreo iniciado correctamente"}
    else:
        raise HTTPException(status_code=400, detail="No se pudo iniciar el monitoreo")

@router.post("/monitoring/stop")
async def stop_monitoring():
    """Detener el monitoreo automático"""
    success = monitoring_scheduler.stop_monitoring()
    if success:
        return {"message": "Monitoreo detenido correctamente"}
    else:
        raise HTTPException(status_code=400, detail="No se pudo detener el monitoreo")

@router.post("/monitoring/manual-iteration")
async def execute_manual_iteration():
    """Ejecutar una iteración manual de monitoreo"""
    result = monitoring_scheduler.execute_manual_iteration()
    if result["success"]:
        return result
    else:
        raise HTTPException(status_code=500, detail=result["error"])

@router.post("/monitoring/continuous/enable")
async def enable_continuous_mode(delay_seconds: int = 10):
    """🚀 Activar modo continuo de ejecución para pruebas"""
    result = monitoring_scheduler.enable_continuous_mode(delay_seconds)
    return result

@router.post("/monitoring/continuous/disable")
async def disable_continuous_mode():
    """🛑 Desactivar modo continuo de ejecución"""
    result = monitoring_scheduler.disable_continuous_mode()
    return result

@router.get("/logs")
async def get_system_logs(
    level: Optional[str] = Query(None, description="Filtrar por nivel de log"),
    module: Optional[str] = Query(None, description="Filtrar por módulo"),
    limit: int = Query(100, description="Límite de registros"),
    db: Session = Depends(get_db)
):
    """Obtener logs del sistema"""
    query = db.query(SystemLog)
    
    if level:
        query = query.filter(SystemLog.level == level)
    
    if module:
        query = query.filter(SystemLog.module == module)
    
    logs = query.order_by(SystemLog.timestamp.desc()).limit(limit).all()
    
    return [
        {
            "id": log.id,
            "level": log.level,
            "message": log.message,
            "module": log.module,
            "timestamp": log.timestamp.isoformat(),
            "session_id": log.session_id
        }
        for log in logs
    ]

@router.get("/stats/daily")
async def get_daily_stats(
    days: int = Query(7, description="Número de días"),
    db: Session = Depends(get_db)
):
    """Obtener estadísticas diarias"""
    from datetime import timedelta
    
    end_date = date.today()
    start_date = end_date - timedelta(days=days-1)
    
    # Estadísticas por día
    daily_stats = []
    current_date = start_date
    
    while current_date <= end_date:
        day_str = current_date.isoformat()
        
        # Contar agencias monitoreadas
        agencies_count = db.query(SalesRecord.agency_code).filter(
            SalesRecord.capture_day == day_str
        ).distinct().count()
        
        # Contar alertas generadas
        alerts_count = db.query(Alert).filter(Alert.alert_day == day_str).count()
        
        # Contar alertas reportadas
        reported_alerts = db.query(Alert).filter(
            Alert.alert_day == day_str,
            Alert.is_reported == True
        ).count()
        
        # Calcular ventas totales
        total_sales = db.query(func.sum(SalesRecord.sales)).filter(
            SalesRecord.capture_day == day_str
        ).scalar() or 0
        
        daily_stats.append({
            "date": day_str,
            "agencies_monitored": agencies_count,
            "alerts_generated": alerts_count,
            "alerts_reported": reported_alerts,
            "total_sales": float(total_sales)
        })
        
        current_date += timedelta(days=1)
    
    return {
        "period": f"{start_date} to {end_date}",
        "daily_stats": daily_stats
    }

@router.get("/health")
async def health_check():
    """Endpoint de verificación de salud"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "monitoring_active": monitoring_scheduler.is_running
    }

# Modelos para configuración
class MonitoringSettings(BaseModel):
    monitoringInterval: int
    browserHeadless: bool
    filterSuriel: bool
    filterTotalGeneral: bool
    enableGrowthAlerts: bool
    enableThresholdAlerts: bool
    salesThreshold: int
    balanceThreshold: int
    growthVariation: int
    sustainedGrowth: int

@router.post("/settings/update")
async def update_settings(settings: MonitoringSettings):
    """Actualizar configuración del sistema"""
    try:
        # Validaciones
        if settings.monitoringInterval < 10:
            raise HTTPException(status_code=400, detail="El intervalo mínimo es de 10 minutos")
        
        if settings.salesThreshold < 1000:
            raise HTTPException(status_code=400, detail="El umbral de ventas mínimo es $1,000")
            
        if settings.balanceThreshold < 1000:
            raise HTTPException(status_code=400, detail="El umbral de balance mínimo es $1,000")
        
        # Actualizar configuración en el scheduler
        monitoring_scheduler.update_settings(settings.dict())
        
        return {
            "message": "Configuración actualizada correctamente",
            "settings": settings.dict()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error actualizando configuración: {str(e)}")

@router.get("/settings")
async def get_settings():
    """Obtener configuración actual del sistema"""
    try:
        current_settings = monitoring_scheduler.get_settings()
        return current_settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo configuración: {str(e)}")

# ========================================
# 🧠 RUTAS DE INTELIGENCIA ARTIFICIAL
# ========================================

@router.get("/intelligence/status", response_model=IntelligenceStatusResponse)
async def get_intelligence_status():
    """Obtener estado completo del sistema de inteligencia artificial"""
    try:
        status = monitoring_scheduler.get_intelligence_status()
        return IntelligenceStatusResponse(**status)
    except Exception as e:
        return IntelligenceStatusResponse(
            intelligence_enabled=False,
            error=f"Error obteniendo estado: {str(e)}"
        )

@router.post("/intelligence/toggle")
async def toggle_intelligence_system(request: IntelligenceToggleRequest):
    """Activar/desactivar el sistema de inteligencia artificial"""
    try:
        monitoring_scheduler.toggle_intelligence(request.enabled)
        status = "activado" if request.enabled else "desactivado"
        return {
            "message": f"Sistema de inteligencia {status} correctamente",
            "enabled": request.enabled
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error cambiando estado: {str(e)}")

@router.get("/intelligence/predictions", response_model=PredictionResponse)
async def get_failure_predictions():
    """Obtener predicciones de riesgo de fallo para próxima iteración"""
    try:
        current_metrics = {
            'current_hour': datetime.now().hour,
            'day_of_week': datetime.now().weekday()
        }
        
        prediction = intelligence_engine.predict_failure_risk(current_metrics)
        
        return PredictionResponse(
            prediction_type=prediction.prediction_type,
            probability=prediction.probability,
            confidence=prediction.confidence,
            recommendation=prediction.recommendation,
            details=prediction.details
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo predicciones: {str(e)}")

@router.get("/intelligence/anomalies")
async def get_anomaly_detection():
    """Detectar anomalías en el comportamiento del sistema"""
    try:
        current_metrics = {
            'current_hour': datetime.now().hour,
            'day_of_week': datetime.now().weekday()
        }
        
        anomalies = intelligence_engine.detect_anomalies(current_metrics)
        
        return {
            "timestamp": datetime.now().isoformat(),
            "anomalies_detected": len(anomalies),
            "anomalies": anomalies
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error detectando anomalías: {str(e)}")

@router.get("/intelligence/optimizations", response_model=List[OptimizationResponse])
async def get_optimization_recommendations():
    """Obtener recomendaciones de optimización del sistema"""
    try:
        recommendations = intelligence_engine.optimize_parameters()
        
        return [
            OptimizationResponse(
                parameter=rec.parameter,
                current_value=str(rec.current_value),
                recommended_value=str(rec.recommended_value),
                confidence=rec.confidence,
                reason=rec.reason,
                expected_improvement=rec.expected_improvement
            )
            for rec in recommendations
        ]
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo optimizaciones: {str(e)}")

@router.post("/intelligence/apply-optimization")
async def apply_optimization(parameter: str, value: str):
    """Aplicar manualmente una optimización específica"""
    try:
        # Validar parámetros permitidos
        allowed_params = [
            "monitoring_interval", "sales_threshold", "page_load_wait",
            "table_load_wait", "element_click_wait"
        ]
        
        if parameter not in allowed_params:
            raise HTTPException(
                status_code=400, 
                detail=f"Parámetro no permitido. Permitidos: {allowed_params}"
            )
        
        # Aplicar optimización según el tipo
        if parameter == "monitoring_interval":
            interval = int(value)
            if interval < 10 or interval > 60:
                raise HTTPException(status_code=400, detail="Intervalo debe estar entre 10 y 60 minutos")
            monitoring_scheduler._update_monitoring_interval(interval)
            
        elif parameter == "sales_threshold":
            threshold = float(value)
            if threshold < 1000 or threshold > 100000:
                raise HTTPException(status_code=400, detail="Umbral debe estar entre $1,000 y $100,000")
            monitoring_scheduler._update_sales_threshold(threshold)
            
        elif parameter.endswith("_wait"):
            wait_time = float(value)
            if wait_time < 1 or wait_time > 30:
                raise HTTPException(status_code=400, detail="Tiempo de espera debe estar entre 1 y 30 segundos")
            wait_type = parameter.replace("_wait", "")
            monitoring_scheduler._update_wait_time(wait_type, wait_time)
        
        return {
            "message": f"Optimización aplicada: {parameter} = {value}",
            "parameter": parameter,
            "value": value
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error aplicando optimización: {str(e)}")

@router.get("/intelligence/adaptive-config")
async def get_adaptive_configuration():
    """Obtener configuración adaptiva actual del sistema"""
    try:
        config = intelligence_engine.get_adaptive_config()
        return {
            "timestamp": datetime.now().isoformat(),
            "adaptive_config": config
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo configuración: {str(e)}")

@router.post("/intelligence/update-config")
async def update_adaptive_config(request: AdaptiveConfigRequest):
    """Actualizar configuración adaptiva del sistema"""
    try:
        intelligence_engine.update_adaptive_config(request.config_updates)
        return {
            "message": "Configuración adaptiva actualizada",
            "updates": request.config_updates
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error actualizando configuración: {str(e)}")

@router.get("/intelligence/metrics/history")
async def get_system_metrics_history(
    days: int = Query(7, description="Días de historial"),
    db: Session = Depends(get_db)
):
    """Obtener historial de métricas del sistema para análisis"""
    try:
        from datetime import timedelta
        
        # Consultar métricas históricas
        start_date = datetime.now() - timedelta(days=days)
        
        # Query directo a la base de datos
        query = """
        SELECT 
            timestamp,
            iteration_success,
            iteration_duration,
            records_obtained,
            alerts_generated,
            error_type,
            website_response_time,
            memory_usage,
            cpu_usage
        FROM system_metrics 
        WHERE timestamp >= ?
        ORDER BY timestamp DESC
        LIMIT 200
        """
        
        # Para este endpoint, necesitamos acceso directo a SQLite
        import sqlite3
        conn = sqlite3.connect("monitoring.db")
        cursor = conn.cursor()
        
        cursor.execute(query, (start_date.isoformat(),))
        rows = cursor.fetchall()
        conn.close()
        
        # Formatear datos
        metrics_history = []
        for row in rows:
            metrics_history.append({
                "timestamp": row[0],
                "iteration_success": bool(row[1]),
                "iteration_duration": row[2],
                "records_obtained": row[3],
                "alerts_generated": row[4],
                "error_type": row[5],
                "website_response_time": row[6],
                "memory_usage": row[7],
                "cpu_usage": row[8]
            })
        
        # Calcular estadísticas
        if metrics_history:
            successful_iterations = [m for m in metrics_history if m["iteration_success"]]
            
            stats = {
                "total_iterations": len(metrics_history),
                "successful_iterations": len(successful_iterations),
                "success_rate": len(successful_iterations) / len(metrics_history) * 100,
                "avg_duration": sum(m["iteration_duration"] or 0 for m in successful_iterations) / len(successful_iterations) if successful_iterations else 0,
                "avg_records": sum(m["records_obtained"] or 0 for m in successful_iterations) / len(successful_iterations) if successful_iterations else 0,
                "total_alerts": sum(m["alerts_generated"] or 0 for m in metrics_history),
            }
        else:
            stats = {
                "total_iterations": 0,
                "successful_iterations": 0,
                "success_rate": 0,
                "avg_duration": 0,
                "avg_records": 0,
                "total_alerts": 0
            }
        
        return {
            "period": f"Últimos {days} días",
            "statistics": stats,
            "metrics_history": metrics_history
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo métricas: {str(e)}")

@router.get("/intelligence/learning-insights")
async def get_learning_insights():
    """Obtener insights del aprendizaje automático del sistema"""
    try:
        # Obtener datos de configuración adaptiva
        adaptive_config = intelligence_engine.get_adaptive_config()
        
        # Obtener predicciones actuales
        current_metrics = {
            'current_hour': datetime.now().hour,
            'day_of_week': datetime.now().weekday()
        }
        
        failure_prediction = intelligence_engine.predict_failure_risk(current_metrics)
        anomalies = intelligence_engine.detect_anomalies(current_metrics)
        optimizations = intelligence_engine.optimize_parameters()
        
        # Calcular insights
        insights = {
            "learning_status": {
                "models_trained": True,  # Simplificado por ahora
                "data_points_available": "Suficientes",
                "last_training": "Continuo",
                "confidence_level": "Alto"
            },
            "performance_trends": {
                "success_rate_trend": "Estable",
                "response_time_trend": "Mejorando",
                "optimization_effectiveness": "Alta"
            },
            "adaptive_behaviors": {
                "timeouts_adjusted": len([k for k in adaptive_config.get("wait_times", {}).keys()]),
                "thresholds_optimized": len([k for k in adaptive_config.get("thresholds", {}).keys()]),
                "intervals_adapted": 1 if "monitoring_interval" in adaptive_config else 0
            },
            "risk_assessment": {
                "current_risk_level": failure_prediction.details.get("risk_level", "Desconocido"),
                "risk_factors": failure_prediction.details.get("main_factors", []),
                "mitigation_actions": len([o for o in optimizations if o.confidence > 0.7])
            },
            "recommendations": {
                "high_confidence": [o for o in optimizations if o.confidence > 0.8],
                "medium_confidence": [o for o in optimizations if 0.6 <= o.confidence <= 0.8],
                "anomalies_to_investigate": anomalies
            }
        }
        
        return {
            "timestamp": datetime.now().isoformat(),
            "insights": insights
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo insights: {str(e)}")

# ========================================
# 🔍 RUTAS DE DOM INTELLIGENCE
# ========================================

@router.get("/dom-intelligence/status")
async def get_dom_intelligence_status():
    """Obtener estado del sistema de observación DOM"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        performance_report = dom_intelligence.generate_performance_report()
        
        return {
            "timestamp": datetime.now().isoformat(),
            "learning_active": dom_intelligence.learning_active,
            "interactions_recorded": len(dom_intelligence.interaction_buffer),
            "current_optimizations": len(dom_intelligence.optimizer.current_optimizations),
            "performance_stats": performance_report['stats_24h'],
            "problem_elements": performance_report['problem_elements'],
            "recommendations_available": len(performance_report['recommendations']),
            "anomalies_detected": len(performance_report['anomalies'])
        }
        
    except Exception as e:
        return {
            "timestamp": datetime.now().isoformat(),
            "learning_active": False,
            "error": f"DOM Intelligence no disponible: {str(e)}"
        }

@router.get("/dom-intelligence/performance-report")
async def get_dom_performance_report():
    """Obtener reporte completo de rendimiento DOM"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        report = dom_intelligence.generate_performance_report()
        
        return {
            "timestamp": datetime.now().isoformat(),
            "report": report
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando reporte: {str(e)}")

@router.get("/dom-intelligence/interactions")
async def get_dom_interactions(
    selector: Optional[str] = Query(None, description="Filtrar por selector"),
    action_type: Optional[str] = Query(None, description="Filtrar por tipo de acción"),
    hours_back: int = Query(24, description="Horas hacia atrás"),
    limit: int = Query(100, description="Límite de resultados")
):
    """Obtener historial de interacciones DOM"""
    try:
        from ..dom_intelligence import dom_intelligence
        import sqlite3
        from datetime import timedelta
        
        conn = sqlite3.connect(dom_intelligence.db.db_path)
        
        # Construir query
        query = """
            SELECT timestamp, action_type, selector, duration, success, 
                   element_found, page_url, context, error_message
            FROM dom_interactions 
            WHERE timestamp > datetime('now', '-{} hours')
        """.format(hours_back)
        
        params = []
        
        if selector:
            query += " AND selector LIKE ?"
            params.append(f"%{selector}%")
        
        if action_type:
            query += " AND action_type = ?"
            params.append(action_type)
        
        query += " ORDER BY timestamp DESC LIMIT ?"
        params.append(limit)
        
        cursor = conn.cursor()
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        interactions = []
        for row in rows:
            interactions.append({
                "timestamp": row[0],
                "action_type": row[1],
                "selector": row[2],
                "duration": row[3],
                "success": bool(row[4]),
                "element_found": bool(row[5]),
                "page_url": row[6],
                "context": row[7],
                "error_message": row[8]
            })
        
        return {
            "total_found": len(interactions),
            "filters_applied": {
                "selector": selector,
                "action_type": action_type,
                "hours_back": hours_back
            },
            "interactions": interactions
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo interacciones: {str(e)}")

@router.get("/dom-intelligence/element-performance/{selector}")
async def get_element_performance(selector: str):
    """Obtener análisis de rendimiento de un elemento específico"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        performance = dom_intelligence.analyzer.analyze_element_performance(selector)
        
        if performance:
            return {
                "selector": performance.selector,
                "metrics": {
                    "total_interactions": performance.total_interactions,
                    "success_rate": performance.success_rate,
                    "avg_duration": performance.avg_duration,
                    "min_duration": performance.min_duration,
                    "max_duration": performance.max_duration,
                    "optimal_timeout": performance.optimal_timeout
                },
                "failure_patterns": performance.failure_patterns,
                "context_performance": performance.context_performance,
                "analysis_timestamp": datetime.now().isoformat()
            }
        else:
            return {
                "selector": selector,
                "error": "No hay datos suficientes para este elemento"
            }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analizando elemento: {str(e)}")

@router.get("/dom-intelligence/anomalies")
async def get_dom_anomalies(days_back: int = Query(7, description="Días hacia atrás")):
    """Detectar anomalías en el rendimiento DOM"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        anomalies = dom_intelligence.analyzer.detect_performance_anomalies(days_back)
        
        return {
            "timestamp": datetime.now().isoformat(),
            "detection_period": f"Últimos {days_back} días",
            "anomalies_found": len(anomalies),
            "anomalies": anomalies
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error detectando anomalías: {str(e)}")

@router.get("/dom-intelligence/optimizations")
async def get_dom_optimizations():
    """Obtener recomendaciones de optimización DOM"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        recommendations = dom_intelligence.optimizer.generate_recommendations()
        
        return {
            "timestamp": datetime.now().isoformat(),
            "recommendations_available": len(recommendations),
            "recommendations": [
                {
                    "optimization_type": rec.optimization_type,
                    "current_value": rec.current_value,
                    "recommended_value": rec.recommended_value,
                    "confidence": rec.confidence,
                    "expected_improvement": rec.expected_improvement,
                    "reason": rec.reason,
                    "affected_elements": rec.affected_elements
                }
                for rec in recommendations
            ]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo optimizaciones: {str(e)}")

@router.post("/dom-intelligence/apply-optimization")
async def apply_dom_optimization(
    optimization_type: str,
    target_element: str,
    new_value: str
):
    """Aplicar una optimización DOM específica"""
    try:
        from ..dom_intelligence import dom_intelligence, OptimizationRecommendation
        
        # Crear recomendación para aplicar
        recommendation = OptimizationRecommendation(
            optimization_type=optimization_type,
            current_value="auto_detected",
            recommended_value=new_value,
            confidence=1.0,  # Manual application has full confidence
            expected_improvement="Mejora manual aplicada",
            reason="Optimización aplicada manualmente por el usuario",
            affected_elements=[target_element]
        )
        
        success = dom_intelligence.optimizer.apply_optimization(recommendation)
        
        if success:
            return {
                "message": "Optimización aplicada correctamente",
                "optimization_type": optimization_type,
                "target_element": target_element,
                "new_value": new_value,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=400, detail="No se pudo aplicar la optimización")
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error aplicando optimización: {str(e)}")

@router.post("/dom-intelligence/reset-optimizations")
async def reset_dom_optimizations():
    """Resetear todas las optimizaciones DOM aplicadas"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        dom_intelligence.reset_optimizations()
        
        return {
            "message": "Todas las optimizaciones DOM han sido reseteadas",
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reseteando optimizaciones: {str(e)}")

@router.post("/dom-intelligence/toggle-learning")
async def toggle_dom_learning():
    """Activar/desactivar el aprendizaje DOM"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        new_status = dom_intelligence.toggle_learning()
        
        return {
            "message": f"Aprendizaje DOM {'activado' if new_status else 'desactivado'}",
            "learning_active": new_status,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error cambiando estado de aprendizaje: {str(e)}")

@router.get("/dom-intelligence/optimal-timeouts")
async def get_optimal_timeouts():
    """Obtener timeouts óptimos predichos por IA"""
    try:
        from ..dom_intelligence import dom_intelligence
        
        optimal_timeouts = dom_intelligence.analyzer.predict_optimal_timeouts()
        
        return {
            "timestamp": datetime.now().isoformat(),
            "total_elements_analyzed": len(optimal_timeouts),
            "optimal_timeouts": optimal_timeouts
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo timeouts óptimos: {str(e)}")

@router.get("/dom-intelligence/stats")
async def get_dom_intelligence_stats():
    """Obtener estadísticas generales del sistema DOM Intelligence"""
    try:
        from ..dom_intelligence import dom_intelligence
        import sqlite3
        
        conn = sqlite3.connect(dom_intelligence.db.db_path)
        
        # Estadísticas generales
        stats_query = """
            SELECT 
                COUNT(*) as total_interactions,
                COUNT(DISTINCT selector) as unique_selectors,
                AVG(duration) as avg_duration,
                AVG(CAST(success AS FLOAT)) as success_rate,
                COUNT(DISTINCT context) as unique_contexts
            FROM dom_interactions
            WHERE timestamp > datetime('now', '-24 hours')
        """
        
        cursor = conn.cursor()
        cursor.execute(stats_query)
        general_stats = cursor.fetchone()
        
        # Top selectores problemáticos
        problem_selectors_query = """
            SELECT selector, 
                   COUNT(*) as interactions,
                   AVG(CAST(success AS FLOAT)) as success_rate,
                   AVG(duration) as avg_duration
            FROM dom_interactions 
            WHERE timestamp > datetime('now', '-7 days')
            GROUP BY selector
            HAVING interactions >= 3
            ORDER BY success_rate ASC, avg_duration DESC
            LIMIT 5
        """
        
        cursor.execute(problem_selectors_query)
        problem_selectors = cursor.fetchall()
        
        # Estadísticas por tipo de acción
        action_stats_query = """
            SELECT action_type, 
                   COUNT(*) as count,
                   AVG(duration) as avg_duration,
                   AVG(CAST(success AS FLOAT)) as success_rate
            FROM dom_interactions 
            WHERE timestamp > datetime('now', '-24 hours')
            GROUP BY action_type
            ORDER BY count DESC
        """
        
        cursor.execute(action_stats_query)
        action_stats = cursor.fetchall()
        
        conn.close()
        
        return {
            "timestamp": datetime.now().isoformat(),
            "period": "Últimas 24 horas",
            "general_stats": {
                "total_interactions": general_stats[0] or 0,
                "unique_selectors": general_stats[1] or 0,
                "avg_duration": round(general_stats[2] or 0, 3),
                "success_rate": round((general_stats[3] or 0) * 100, 2),
                "unique_contexts": general_stats[4] or 0
            },
            "problem_selectors": [
                {
                    "selector": row[0],
                    "interactions": row[1],
                    "success_rate": round(row[2] * 100, 2),
                    "avg_duration": round(row[3], 3)
                }
                for row in problem_selectors
            ],
            "action_stats": [
                {
                    "action_type": row[0],
                    "count": row[1],
                    "avg_duration": round(row[2], 3),
                    "success_rate": round(row[3] * 100, 2)
                }
                for row in action_stats
            ],
            "current_optimizations": len(dom_intelligence.optimizer.current_optimizations),
            "learning_status": "Activo" if dom_intelligence.learning_active else "Pausado"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo estadísticas DOM: {str(e)}")

# ========================================
# 🔗 RUTAS DE SISTEMA HÍBRIDO
# ========================================

# Importar scheduler híbrido
try:
    from ..scheduler_hybrid import monitoring_scheduler as hybrid_scheduler
    HYBRID_SCHEDULER_AVAILABLE = True
except ImportError:
    hybrid_scheduler = monitoring_scheduler  # Fallback al scheduler clásico
    HYBRID_SCHEDULER_AVAILABLE = False

@router.get("/hybrid/status")
async def get_hybrid_status():
    """Obtener estado del sistema híbrido"""
    if HYBRID_SCHEDULER_AVAILABLE:
        return hybrid_scheduler.get_status()
    else:
        return {
            "error": "Sistema híbrido no disponible",
            "using_classic_scheduler": True,
            "classic_status": monitoring_scheduler.get_status()
        }

@router.post("/hybrid/enable-intelligent")
async def enable_intelligent_system(percentage: float = 100.0):
    """Habilitar sistema inteligente gradualmente"""
    if not HYBRID_SCHEDULER_AVAILABLE:
        raise HTTPException(status_code=404, detail="Sistema híbrido no disponible")
    
    success = hybrid_scheduler.enable_intelligent_system(percentage)
    return {"success": success, "percentage": percentage}

@router.post("/hybrid/disable-intelligent")
async def disable_intelligent_system():
    """Deshabilitar sistema inteligente"""
    if not HYBRID_SCHEDULER_AVAILABLE:
        raise HTTPException(status_code=404, detail="Sistema híbrido no disponible")
    
    hybrid_scheduler.disable_intelligent_system()
    return {"success": True}

@router.post("/hybrid/comparison-mode")
async def toggle_comparison_mode(enabled: bool):
    """Activar/desactivar modo comparación"""
    if not HYBRID_SCHEDULER_AVAILABLE:
        raise HTTPException(status_code=404, detail="Sistema híbrido no disponible")
    
    if enabled:
        success = hybrid_scheduler.enable_comparison_mode()
    else:
        hybrid_scheduler.disable_comparison_mode()
        success = True
    return {"success": success, "comparison_mode": enabled}

@router.get("/hybrid/performance-comparison")
async def get_performance_comparison():
    """Obtener comparación de rendimiento entre sistemas"""
    if not HYBRID_SCHEDULER_AVAILABLE:
        raise HTTPException(status_code=404, detail="Sistema híbrido no disponible")
    
    return hybrid_scheduler.get_performance_comparison()

@router.get("/hybrid/availability")
async def check_hybrid_availability():
    """Verificar disponibilidad del sistema híbrido"""
    return {
        "hybrid_available": HYBRID_SCHEDULER_AVAILABLE,
        "intelligent_system_available": HYBRID_SCHEDULER_AVAILABLE and hasattr(hybrid_scheduler, 'intelligent_scraper') and hybrid_scheduler.intelligent_scraper is not None,
        "message": "Sistema híbrido disponible" if HYBRID_SCHEDULER_AVAILABLE else "Usando sistema clásico"
    }

# 🧠 RUTAS DE ANÁLISIS DE NORMALIDAD DE AGENCIAS
@router.get("/analysis/agency/{agency_code}/normality")
async def analyze_agency_normality(
    agency_code: str,
    current_sales: float,
    lottery_type: str = "RULETA_EXPRESS"
):
    """
    🎯 Análisis rápido de normalidad de una agencia
    Respuesta instantánea sobre si el comportamiento es inusual
    """
    try:
        result = agency_analyzer.quick_analysis(agency_code, current_sales, lottery_type)
        
        return {
            "success": True,
            "agency_code": agency_code,
            "current_sales": float(current_sales),
            "lottery_type": lottery_type,
            "analysis": {
                "is_unusual": bool(result.is_unusual),
                "confidence": float(result.confidence),
                "explanation": result.explanation,
                "recommendation": result.recommendation
            },
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error analizando normalidad de agencia {agency_code}: {e}")
        return {
            "success": False,
            "error": str(e),
            "agency_code": agency_code
        }

@router.get("/analysis/agency/{agency_code}/growth")
async def analyze_agency_growth(
    agency_code: str,
    lottery_type: str = "RULETA_EXPRESS"
):
    """
    📈 Análisis de patrón de crecimiento de una agencia
    """
    try:
        growth_analysis = agency_analyzer.analyze_growth_normality(agency_code, lottery_type)
        
        return {
            "success": True,
            "agency_code": agency_code,
            "lottery_type": lottery_type,
            "growth_analysis": growth_analysis,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error analizando crecimiento de agencia {agency_code}: {e}")
        return {
            "success": False,
            "error": str(e),
            "agency_code": agency_code
        }

@router.post("/analysis/batch-normality")
async def batch_analyze_normality(request: Dict):
    """
    🚀 Análisis de normalidad por lotes
    Analiza múltiples agencias simultáneamente
    """
    try:
        agencies = request.get("agencies", [])
        lottery_type = request.get("lottery_type", "RULETA_EXPRESS")
        
        results = []
        for agency_data in agencies:
            agency_code = agency_data.get("agency_code")
            current_sales = agency_data.get("current_sales", 0)
            
            if agency_code:
                result = agency_analyzer.quick_analysis(agency_code, current_sales, lottery_type)
                results.append({
                    "agency_code": agency_code,
                    "current_sales": current_sales,
                    "is_unusual": result.is_unusual,
                    "confidence": result.confidence,
                    "explanation": result.explanation,
                    "recommendation": result.recommendation
                })
        
        # Estadísticas del lote
        total_agencies = len(results)
        unusual_agencies = len([r for r in results if r["is_unusual"]])
        avg_confidence = float(np.mean([r["confidence"] for r in results])) if results else 0.0
        
        return {
            "success": True,
            "lottery_type": lottery_type,
            "batch_summary": {
                "total_agencies": total_agencies,
                "unusual_agencies": unusual_agencies,
                "unusual_percentage": float((unusual_agencies / total_agencies * 100)) if total_agencies > 0 else 0.0,
                "average_confidence": avg_confidence
            },
            "agency_results": results,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error en análisis por lotes: {e}")
        return {
            "success": False,
            "error": str(e)
        }

# 🧠 === RUTAS DOM LEARNING ENGINE ===

@router.get("/api/v1/dom-learning/table-anomalies")
async def get_table_anomalies():
    """🚨 Detectar anomalías en carga de tablas (lentitud, timeouts)"""
    try:
        anomalies = dom_learner.detect_table_loading_anomalies()
        
        return {
            "success": True,
            "anomalies_count": len(anomalies),
            "anomalies": [
                {
                    "type": a.type,
                    "severity": a.severity,
                    "description": a.description,
                    "affected_selector": a.affected_selector,
                    "baseline_value": a.baseline_value,
                    "current_value": a.current_value,
                    "confidence": a.confidence,
                    "recommendation": a.recommendation,
                    "timestamp": a.timestamp.isoformat()
                }
                for a in anomalies
            ]
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo anomalías de tabla: {e}")
        return {"success": False, "error": str(e)}

@router.post("/api/v1/dom-learning/analyze-iteration")
async def analyze_iteration_performance(data: dict):
    """📊 Analizar rendimiento de iteración específica"""
    try:
        duration = data.get('duration', 0)
        agencies_processed = data.get('agencies_processed', 0)
        
        analysis = dom_learner.analyze_current_iteration_performance(duration, agencies_processed)
        
        return {
            "success": True,
            "analysis": analysis
        }
        
    except Exception as e:
        logger.error(f"Error analizando rendimiento: {e}")
        return {"success": False, "error": str(e)}

@router.get("/api/v1/dom-learning/performance-insights")
async def get_performance_insights():
    """💡 Obtener insights y recomendaciones de optimización"""
    try:
        # Obtener datos de rendimiento recientes
        conn = sqlite3.connect("dom_intelligence.db")
        
        # Estadísticas de últimas 24 horas
        query = """
            SELECT 
                COUNT(*) as total_interactions,
                AVG(duration) as avg_duration,
                AVG(success) * 100 as success_rate,
                COUNT(DISTINCT selector) as unique_selectors
            FROM dom_interactions 
            WHERE timestamp >= datetime('now', '-24 hours')
        """
        
        cursor = conn.execute(query)
        stats = cursor.fetchone()
        conn.close()
        
        # Estadísticas básicas
        basic_stats = {
            "total_interactions": stats[0] if stats[0] else 0,
            "avg_duration": round(stats[1], 2) if stats[1] else 0,
            "success_rate": round(stats[2], 1) if stats[2] else 0,
            "unique_selectors": stats[3] if stats[3] else 0
        }
        
        # Detectar anomalías
        table_anomalies = dom_learner.detect_table_loading_anomalies()
        
        # Generar insights
        insights = []
        
        if basic_stats["success_rate"] < 90:
            insights.append({
                "type": "reliability",
                "severity": "high",
                "title": "Baja tasa de éxito",
                "description": f"Solo {basic_stats['success_rate']}% de interacciones exitosas",
                "recommendation": "Revisar selectores y timeouts"
            })
        
        if basic_stats["avg_duration"] > 5:
            insights.append({
                "type": "performance", 
                "severity": "medium",
                "title": "Interacciones lentas",
                "description": f"Promedio de {basic_stats['avg_duration']}s por interacción",
                "recommendation": "Optimizar selectores más frecuentes"
            })
        
        return {
            "success": True,
            "basic_stats": basic_stats,
            "table_anomalies_count": len(table_anomalies),
            "insights": insights,
            "recommendations": [
                "Monitorear patrones de lentitud en tablas",
                "Implementar timeouts adaptativos",
                "Revisar selectores con baja tasa de éxito"
            ]
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo insights: {e}")
        return {"success": False, "error": str(e)} 